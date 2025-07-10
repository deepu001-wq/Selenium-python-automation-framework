import inspect
import logging

import softest
from openpyxl import Workbook, load_workbook
import csv

class utils(softest.TestCase):
    def assert_list_items(self,list,value):

        for stop in list:
            print("the text is:" + stop.text)
            self.soft_assert(self.assertEqual,stop.text,value)

            if stop.text==value:
                print("test passed")
            else:
                print("test failed")
        self.assert_all()

    def Custom_logger(loglevel=logging.DEBUG):
        # set class/method name from where its called
        logger_name = inspect.stack()[1][3]
        # create logger
        logger = logging.getLogger(logger_name)
        logger.setLevel(loglevel)
        # create console handler or file handler and set the log level

        fh = logging.FileHandler("automation.log")
        # create formatter
        formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s : %(message)s',
                                      datefmt='%d-%m-%Y %H:%M:%S')

        # add formatter to console or filehandler

        fh.setFormatter(formatter)
        # add console handler to logger

        logger.addHandler(fh)
        # application code - log messages
        return logger

    def read_data_from_excel(file_name, sheet):
        datalist = []
        wb = load_workbook(filename=file_name)
        sh = wb[sheet]
        row_ct = sh.max_row
        col_ct = sh.max_column

        for i in range(2, row_ct + 1):
            row = []
            for j in range(1, col_ct + 1):
                row.append(sh.cell(row=i, column=j).value)
            datalist.append(row)
        return datalist

    def read_data_from_csv(file_name):
        #create an empty list
        datalist = []

        #open csv file
        csvdata=open(file_name,"r")
        #create csv reader
        reader=csv.reader(csvdata)

        #skip header
        next(reader)
        #add csv rows to list
        for rows in reader:
            datalist.append(rows)
        return datalist



